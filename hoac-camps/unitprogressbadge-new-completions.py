
from openpyxl import load_workbook, Workbook, worksheet
import sys
import os

if os.environ.get('PROGRESS_FOLDER') == None or os.environ['PROGRESS_FOLDER'] == "":
  sys.exit("Please set the PROGRESS_FOLDER environment variable.") 

if len(sys.argv) < 3:
  sys.exit('Please pass the file names to be compared as arguments.')

file1 = os.environ['PROGRESS_FOLDER'] + "/" + sys.argv[1]
file2 = os.environ['PROGRESS_FOLDER'] + "/" + sys.argv[2]
if len(sys.argv) == 4:
  yesOnly = sys.argv[3] in [ 'Y', 'y' ]
else:
  yesOnly = False

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def dumpRow(r: tuple, yesOnly: bool) -> str:
  retty = []
  if yesOnly:
    r = r[0:2]
  for v in r:
    if v.value == "Yes":
      hld = retty[0]
      retty[0] = f"{bcolors.OKCYAN}{bcolors.BOLD}{hld}"
      retty.append(f"Yes{bcolors.ENDC}")
    else:
      retty.append(v.value)

  return ",".join(retty)

def printTuple(badge: str, r1: tuple, r2: tuple, yesOnly: bool):
  arr1 = []
  arr2 = []
  print(f"{badge}")
  print(f"  {dumpRow(r1, yesOnly)}")
  print(f"  {dumpRow(r2, yesOnly)}")

wb1 = load_workbook(filename = file1)
wb2 = load_workbook(filename = file2)

for ws1 in wb1.worksheets:
  # if ws1.title != "Environmental Science":
  #   continue
  if ws1.title.startswith("Trail To First Class"):
    continue

  # print(f"=== {ws1.title} ================================================================")

  ws2 = wb2[ws1.title]
  # print(f"found: {ws2.title}")
  if ws2 == None:
    sys.exit(1, f"Could not find {ws1.title} in book 2")

  for r1,r2 in zip(ws1,ws2):
    i = 0
    for c1,c2 in zip(r1,r2):
      if i == 2 and yesOnly:
        break
      i += 1
      if c1.value != c2.value:
        # print(f"Difference found!")
        printTuple(ws1.title, r1, r2, yesOnly)
        # print(r1)
        # print(r2)
        break

    # print(z)
  # sys.exit(0)