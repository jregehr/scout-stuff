
from openpyxl import load_workbook

# wb1 = load_workbook(filename = '/Users/regehr/Documents/Personal/Scouts/Bartle/Bartle2023/merit-badges/progress/UnitProgressExport-2023-06-28.xlsx')
wb1 = load_workbook(filename = '/Users/regehr/Documents/Personal/Scouts/Bartle/Bartle2023/merit-badges/progress/UnitProgressExport-2023-06-29.xlsx')
# wb1 = load_workbook(filename = '/Users/regehr/Documents/Personal/Scouts/Bartle/Bartle2023/merit-badges/progress/UnitProgressExport-2023-06-29.xlsx')
wb2 = load_workbook(filename = '/Users/regehr/Documents/Personal/Scouts/Bartle/Bartle2023/merit-badges/progress/UnitProgressExport-2023-07-01_002.xlsx')

ws1 = wb1['Worksheet']
ws2 = wb2['Worksheet']
# ws1.delete_rows(1,2)
# ws1.insert_cols(1)

# ws1["A1"].value = ws1["B1"].value

print(f"ws1 rows: {ws1.max_row}")
print(f"ws2 rows: {ws2.max_row}")